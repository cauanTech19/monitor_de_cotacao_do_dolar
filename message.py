import requests
from requests import Response
import json
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
from typing import TypedDict
import smtplib
from email.mime.text import MIMEText

# --- Configurações ---
EMAIL_REMETENTE: str = "seuemail@gmail.com"
EMAIL_DESTINATARIO: str = "seuemail@gmail.com"  # pode ser outro email se quiser
EMAIL_SENHA_APP: str = "sua_senha_de_app"        # senha de app do Google, não a senha normal

LIMITE_DOLAR: float = 5.80
ARQUIVO_HISTORICO: str = "historico.json"
ARQUIVO_RELATORIO: str = "relatorio.xlsx"

ABAS: dict[str, str] = {
    "historico": "Historico",
    "resumo": "Resumo",
}


class Cotacao(TypedDict):
    """Representa um registro de cotação com data e valor."""

    data: str
    valor: float


class ResultadoRelatorio(TypedDict):
    """Resultado calculado a partir do histórico de cotações."""

    media: float
    maximo: float
    minimo: float
    ultima: float
    tendencia: str


def buscar_cotacao() -> float:
    """Busca a cotação atual do dólar (USD-BRL) via AwesomeAPI.

    Returns:
        Valor de venda do dólar como float.

    Raises:
        requests.exceptions.HTTPError: Se a API retornar status de erro.
        requests.exceptions.Timeout: Se a requisição exceder 10 segundos.
    """
    url: str = "https://economia.awesomeapi.com.br/json/last/USD-BRL"
    response: Response = requests.get(url, timeout=10)
    response.raise_for_status()
    dados: dict = response.json()
    return float(dados["USDBRL"]["bid"])


def salvar_historico(cotacao: float) -> list[Cotacao]:
    """Adiciona a cotação atual ao histórico e persiste em JSON.

    Mantém no máximo os últimos 30 registros para não crescer indefinidamente.

    Args:
        cotacao: Valor do dólar a ser registrado.

    Returns:
        Lista atualizada com todos os registros do histórico.
    """
    historico: list[Cotacao] = []

    if Path(ARQUIVO_HISTORICO).exists():
        with open(ARQUIVO_HISTORICO) as f:
            historico = json.load(f)

    historico.append({
        "data": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "valor": cotacao,
    })

    historico = historico[-30:]

    with open(ARQUIVO_HISTORICO, "w") as f:
        json.dump(historico, f, indent=2)

    return historico


def gerar_relatorio(historico: list[Cotacao]) -> ResultadoRelatorio:
    """Processa o histórico, gera o Excel formatado com gráfico e retorna métricas.

    Cria duas abas no arquivo:
    - Historico: série completa de cotações com gráfico de linha.
    - Resumo: tabela com média, máxima, mínima, última cotação e tendência.

    Args:
        historico: Lista de registros com 'data' e 'valor'.

    Returns:
        Dicionário com as métricas calculadas (media, maximo, minimo, ultima, tendencia).
    """
    df: DataFrame = pd.DataFrame(historico)
    df["data"] = pd.to_datetime(df["data"], format="%Y/%m/%d %H:%M")
    df = df.sort_values("data")

    media: float = round(df["valor"].mean(), 2)
    maximo: float = round(df["valor"].max(), 2)
    minimo: float = round(df["valor"].min(), 2)
    ultima: float = round(float(df["valor"].iloc[-1]), 2)
    tendencia: str = "Alta 📈" if df["valor"].iloc[-1] > df["valor"].iloc[0] else "Baixa 📉"

    resumo: DataFrame = pd.DataFrame({
        "Métrica": ["Última cotação", "Média", "Máxima", "Mínima", "Tendência"],
        "Valor": [
            f"R$ {ultima:.2f}",
            f"R$ {media:.2f}",
            f"R$ {maximo:.2f}",
            f"R$ {minimo:.2f}",
            tendencia,
        ],
    })

    # Salva as abas no Excel
    with pd.ExcelWriter(ARQUIVO_RELATORIO, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=ABAS["historico"], index=False)
        resumo.to_excel(writer, sheet_name=ABAS["resumo"], index=False)

    # Formata cabeçalhos de todas as abas
    wb = load_workbook(ARQUIVO_RELATORIO)

    for nome_aba in ABAS.values():
        ws = wb[nome_aba]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    # Adiciona gráfico de linha na aba Historico
    ws = wb[ABAS["historico"]]
    chart = LineChart()
    chart.title = "Cotação do Dólar"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Data"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    ws.add_chart(chart, "D2")

    wb.save(ARQUIVO_RELATORIO)

    return ResultadoRelatorio(
        media=media,
        maximo=maximo,
        minimo=minimo,
        ultima=ultima,
        tendencia=tendencia,
    )


def enviar_email(mensagem: str) -> None:
    msg: str = MIMEText(mensagem)
    msg['Subject'] = "Relatório Dólar"
    msg['From'] = EMAIL_REMETENTE
    msg['To'] = EMAIL_DESTINATARIO

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(EMAIL_REMETENTE, EMAIL_SENHA_APP)
        server.send_message(msg)



def pipeline() -> None:
    """Executa o pipeline completo de coleta, processamento e notificação.

    Fluxo:
        1. Busca cotação atual via API.
        2. Salva no histórico JSON.
        3. Gera relatório Excel com métricas e gráfico.
        4. Envia resumo formatado via Email.
    """
    try:
        cotacao: float = buscar_cotacao()
        historico: list[Cotacao] = salvar_historico(cotacao)
        resultado: ResultadoRelatorio = gerar_relatorio(historico)

        agora: str = datetime.now().strftime("%Y/%m/%d %H:%M")
        alerta: str = "🚨 Acima do Limite!" if cotacao >= LIMITE_DOLAR else "✅ Dentro do Normal"

        mensagem: str = (
            f"📊 Relatório Dólar — {agora}\n"
            f"━━━━━━━━━━━━━━\n"
            f"Agora:    R$ {resultado['ultima']:.2f}  {alerta}\n"
            f"Média:    R$ {resultado['media']:.2f}\n"
            f"Máxima:   R$ {resultado['maximo']:.2f}\n"
            f"Mínima:   R$ {resultado['minimo']:.2f}\n"
            f"Tendência: {resultado['tendencia']}\n"
            f"━━━━━━━━━━━━━━\n"
            f"Relatório Excel gerado automaticamente."
        )

        enviar_email(mensagem)
        print(f"[{agora}] Pipeline executado. Cotação: R$ {cotacao:.2f}")

    except Exception as e:
        print(f"Erro no pipeline: {e}")


pipeline()
